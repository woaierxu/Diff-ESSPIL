import os
import argparse
import pathlib
import random
import time

import torch
import pdb
import torch.nn as nn
from torch.backends import cudnn
from openpyxl import Workbook, load_workbook

from networks.DiffUVNet import *
from utils.test_3d_patch import *

from networks.Vnet_LA import VNet



os.environ['CUDA_VISIBLE_DEVICES'] = '0'
train_time_step =1000
test_path = f"./experiments_ori/LA/train_la_rncl_lab10/self_train"
root_path = r"/data1/mengqingxu/Dataset/LA/data/"
exp_name = pathlib.Path(test_path).parent.name
num_classes = 2


EXCEL_FILENAME = os.path.join('./experiments/LA', "test_metrics_summary.xlsx")

EXCEL_HEADERS = ['exp_name', 'infer_start_timestep', 'timesteps', 'dice', 'jc', 'hd', 'asd']




with open('./Datasets/la/data_split/test.txt', 'r') as f:
    image_list = f.readlines()
image_list = [root_path+r"/2018LA_Seg_Training Set/" + item.replace('\n', '') + "/mri_norm2.h5" for item in
              image_list]


def create_Vnet(ema=False):
    net = VNet(n_channels=1, n_classes=2, normalization='instancenorm', has_dropout=False)
    net = nn.DataParallel(net)
    model = net.cuda()
    if ema:
        for param in model.parameters():
            param.detach_()
    return model


num_classes_dfu = 2


def creat_DiffUnet_Vencoder(infer_start_timestep=0, ema=False):
    timesteps_value = 20
    net = DiffUnet_predMask_FuseFeature(
        in_channels=1, num_classes=num_classes_dfu, num_filters=(16, 32, 64, 128, 256, 32),
        infer_start_timestep=infer_start_timestep, timesteps=timesteps_value,
        train_time_step=train_time_step
    )
    net = nn.DataParallel(net)
    model = net.cuda()
    if ema:
        for param in model.parameters():
            param.detach_()
    return model, timesteps_value

def init_excel_file():
    """检查Excel文件是否存在，如果不存在则创建并写入表头。"""
    if not os.path.exists(EXCEL_FILENAME):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Metrics"
        sheet.append(EXCEL_HEADERS)
        workbook.save(EXCEL_FILENAME)
        print(f"Created new Excel file: {EXCEL_FILENAME}")

def append_to_excel(data_dict):
    """将一行数据追加到Excel文件中。"""
    try:
        workbook = load_workbook(EXCEL_FILENAME)
        sheet = workbook.active
        row_to_append = [data_dict.get(header, "") for header in EXCEL_HEADERS]
        sheet.append(row_to_append)
        workbook.save(EXCEL_FILENAME)
    except FileNotFoundError:
        print(f"Error: Excel file {EXCEL_FILENAME} not found. Please initialize it first.")
    except Exception as e:
        print(f"An error occurred while writing to Excel: {e}")


def testLA(infer_start_timestep=0):
    net1 = create_Vnet()
    net2, timesteps = creat_DiffUnet_Vencoder(infer_start_timestep)

    model_path1 = os.path.join(test_path, 'best_model.pth')
    net1.load_state_dict(torch.load(str(model_path1)))

    model_path2 = os.path.join(test_path, 'best_model_diffu.pth')
    net2.load_state_dict(torch.load(str(model_path2)))

    net1.eval()
    net2.eval()

    time0 = time.time()
    avg_metric2 = test_all_case_DiffUV_parallel(net1, net2, num_classes=num_classes, image_list=image_list,
                                       patch_size=(112, 112, 80), stride_xy=18, stride_z=4,
                                       model_choose=1, has_feature=True, multi_step=True,num_workers=1,
                                       fuse_feature1=True,nms=1)
    print(f'total use time for infer_start_timestep {infer_start_timestep}: {round((time.time()-time0),4)}s')



    print("diff-UVnet metrics:")
    print(avg_metric2)


    diff_uvnet_data = {
        'exp_name': pathlib.Path(test_path).parent.name,
        'infer_start_timestep': infer_start_timestep,
        'timesteps': timesteps,
        'dice': avg_metric2[0],
        'jc': avg_metric2[1],
        'hd': avg_metric2[2],
        'asd': avg_metric2[3],

    }


    append_to_excel(diff_uvnet_data)
    print(f"Results for infer_start_timestep={infer_start_timestep} have been saved to {EXCEL_FILENAME}")



if __name__ == '__main__':
    cudnn.benchmark = False
    cudnn.deterministic = True
    torch.manual_seed(1345)
    torch.cuda.manual_seed(1345)
    random.seed(1345)
    np.random.seed(1345)
    init_excel_file()

    for i in range(1):
        current_infer_timestep = 11
        print(f"\n--- Running {exp_name} with infer_start_timestep = {current_infer_timestep} ---")
        testLA(infer_start_timestep=current_infer_timestep)

