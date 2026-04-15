import os
import argparse
import pathlib
import random
import time

import torch
import pdb
import torch.nn as nn
from torch.backends import cudnn
from openpyxl import Workbook, load_workbook

from networks.DiffUVNet import *
from pancreas.test_util import test_all_case_MQX
from utils.test_3d_patch import *

from networks.VNet_EMHSA import VNet



os.environ['CUDA_VISIBLE_DEVICES'] = '0'
test_path = "./pancreas/model/train_pancreas_rncl_lab10/self_train"
root_path = r"/data1/mengqingxu/Dataset/Pancreas/preprocess/"
exp_name = pathlib.Path(test_path).parent.name
num_classes = 2


EXCEL_FILENAME = os.path.join('./pancreas/model', "test_metrics_summary.xlsx")
EXCEL_HEADERS = ['exp_name', 'infer_start_timestep', 'timesteps', 'dice', 'jc', 'hd', 'asd']

with open('./Datasets/pancreas/data_split/test.txt', 'r') as f:
    image_list = f.readlines()
image_list = [root_path + item.replace('\n', '').replace('data','')
              for
              item
              in image_list]

def create_Vnet(ema=False):
    net = VNet(n_channels=1, n_classes=2, normalization='instancenorm', has_dropout=False)
    net = nn.DataParallel(net)
    model = net.cuda()
    if ema:
        for param in model.parameters():
            param.detach_()
    return model




def creat_DiffUnet_Vencoder(infer_start_timestep=0, ema=False):
    timesteps_value = 20
    net = DiffUnet_predMask_FuseFeature(
        in_channels=1, num_classes=num_classes, num_filters=(16, 32, 64, 128, 256, 32),
        infer_start_timestep=infer_start_timestep, timesteps=timesteps_value
    )
    net = nn.DataParallel(net)
    model = net.cuda()
    if ema:
        for param in model.parameters():
            param.detach_()
    return model, timesteps_value





def init_excel_file():
    if not os.path.exists(EXCEL_FILENAME):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Metrics"
        sheet.append(EXCEL_HEADERS)
        workbook.save(EXCEL_FILENAME)
        print(f"Created new Excel file: {EXCEL_FILENAME}")


def append_to_excel(data_dict):
    """将一行数据追加到Excel文件中。"""
    try:
        workbook = load_workbook(EXCEL_FILENAME)
        sheet = workbook.active
        row_to_append = [data_dict.get(header, "") for header in EXCEL_HEADERS]
        sheet.append(row_to_append)
        workbook.save(EXCEL_FILENAME)
    except FileNotFoundError:
        print(f"Error: Excel file {EXCEL_FILENAME} not found. Please initialize it first.")
    except Exception as e:
        print(f"An error occurred while writing to Excel: {e}")


def testLA(infer_start_timestep=0):
    net1 = create_Vnet()
    net2, timesteps = creat_DiffUnet_Vencoder(infer_start_timestep)
    try:
        model_path1 = os.path.join(test_path, 'best_ema_10_self.pth')
        net1_weight = torch.load(str(model_path1))['net']
        net1.load_state_dict(net1_weight)
    except:
        model_path1 = os.path.join(test_path, 'best_ema_20_self.pth')
        net1_weight = torch.load(str(model_path1))['net']
        net1.load_state_dict(net1_weight)


    try:
        model_path2 = os.path.join(test_path, 'best_ema_10_self_diffu.pth')
        net2_weight = torch.load(str(model_path2))['net']
        net2.load_state_dict(net2_weight)
    except:
        model_path2 = os.path.join(test_path, 'best_ema_20_self_diffu.pth')
        net2.load_state_dict(torch.load(str(model_path2))['net'])


    net1.eval()
    net2.eval()
    net2.module.fuse_feature = True
    time0 = time.time()
    avg_metric2, m_list = test_all_case_MQX(net1, net2, image_list, num_classes=num_classes,
                                           patch_size=(96,96,96), stride_xy=16, stride_z=16,
                                           save_result=False, test_save_path='./save', DTC=False, nms=1,
                                           model_choose=1,
                                        )
    print(f'total use time for infer_start_timestep {infer_start_timestep}: {round((time.time()-time0),4)}s')


    print("diff-UVnet metrics:")
    print(avg_metric2)

    diff_uvnet_data = {
        'exp_name': pathlib.Path(test_path).parent.name,
        'infer_start_timestep': infer_start_timestep,
        'timesteps': timesteps,
        'dice': avg_metric2[0],
        'jc': avg_metric2[1],
        'hd': avg_metric2[2],
        'asd': avg_metric2[3],

    }

    append_to_excel(diff_uvnet_data)
    print(f"Results for infer_start_timestep={infer_start_timestep} have been saved to {EXCEL_FILENAME}")


if __name__ == '__main__':
    cudnn.benchmark = False
    cudnn.deterministic = True
    torch.manual_seed(1345)
    torch.cuda.manual_seed(1345)
    random.seed(1345)
    np.random.seed(1345)
    init_excel_file()

    for i in range(19):
        current_infer_timestep = i+1
        print(f"\n--- Running {exp_name} with infer_start_timestep = {current_infer_timestep} ---")
        testLA(infer_start_timestep=current_infer_timestep)


